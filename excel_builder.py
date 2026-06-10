import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ── Palette ────────────────────────────────────────────────────────────────
BLUE_DARK   = "1B3A6B"   # header bg
BLUE_MID    = "2E5FA3"   # sub-header
BLUE_LIGHT  = "D6E4F7"   # alternating row
WHITE       = "FFFFFF"
ACCENT      = "F4A300"   # answer highlight
FONT_HEADER = "FFFFFF"


def _border(style="thin"):
    s = Side(style=style, color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _style_header_row(ws, row_num, ncols, bg=BLUE_DARK, fg=FONT_HEADER, sz=11):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = _header_fill(bg)
        cell.font = Font(bold=True, color=fg, size=sz, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()


def _col_widths():
    return {
        "Title":      40,
        "Question":   60,
        "Option A":   30,
        "Option B":   30,
        "Option C":   30,
        "Option D":   30,
        "Option E":   30,
        "Answer":     10,
        "Solution":   60,
        "Difficulty": 14,
        "Type":       10,
        "URL":        40,
    }


def build_excel(data: dict) -> bytes:
    """
    data = {
      "PS": [rows...],
      "CR": [rows...],
      "RC": [rows...],
    }
    Returns Excel file as bytes.
    Sheet structure:
      - Quantitative  → PS data
      - Verbal → CR + RC data (with Type column to distinguish)
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_map = {
        "Quantitative": ["PS"],
        "Verbal":       ["CR", "RC"],
    }

    widths = _col_widths()
    COLS = list(widths.keys())

    for sheet_name, sections in sheet_map.items():
        # Combine rows for this sheet
        rows = []
        for sec in sections:
            rows.extend(data.get(sec, []))

        ws = wb.create_sheet(title=sheet_name)

        if not rows:
            ws["A1"] = "No data scraped for this section."
            continue

        df = pd.DataFrame(rows, columns=COLS)

        # ── Title banner ──
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"GMAT Club — {sheet_name} Questions"
        title_cell.fill = _header_fill(BLUE_DARK)
        title_cell.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Sub-header (section labels if multiple) ──
        # ── Column headers ──
        for col_idx, col_name in enumerate(COLS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
        _style_header_row(ws, 2, len(COLS), bg=BLUE_MID)
        ws.row_dimensions[2].height = 22

        # ── Data rows ──
        for r_idx, row in enumerate(df.itertuples(index=False), start=3):
            is_even = (r_idx % 2 == 0)
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=str(value) if value else "")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = _border()
                col_name = COLS[c_idx - 1]
                # Alternating row shading
                if is_even:
                    cell.fill = _header_fill(BLUE_LIGHT)
                # Highlight Answer column
                if col_name == "Answer" and str(value) in list("ABCDE"):
                    cell.fill = _header_fill(ACCENT)
                    cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
            ws.row_dimensions[r_idx].height = 60

        # ── Column widths ──
        for col_idx, col_name in enumerate(COLS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 20)

        # ── Freeze panes ──
        ws.freeze_panes = "A3"

        # ── Auto-filter ──
        ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

    # ── Save to bytes ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
